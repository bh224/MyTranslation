from lib2to3.pgen2.parse import ParseError
from django.db import transaction
from rest_framework.test import APITestCase
from projects.models import Project, Project_Member
from users.models import User
from trans.models import Translation

import MySQLdb
import boto3
import io
from openpyxl import Workbook, load_workbook

class TestProjectServices(APITestCase):
    def create_user(self, username, is_admin):
        user = User.objects.create(username=username, is_admin=is_admin)
        user.set_password("1234")
        user.save()
        return user
    
    def test_create_a_project(self) -> None:
        # Given
        manager = self.create_user("manager", True)
        author = self.create_user("author", False)
        checker = self.create_user("checker", False)

        project = Project.objects.create(category="webtoon", title="여신강림", description="한일/네이버웹툰/시즌1", uploader=manager)

        Project_Member.objects.create(project=project, member=manager, role="manager")
        Project_Member.objects.create(project=project, member=author, role="author")
        Project_Member.objects.create(project=project, member=checker, role="checker")

        # When
        members = project.members.all()
        title = author.projects.get().title
        role = Project_Member.objects.filter(project=project, member=checker).get().role

        # Then
        self.assertEqual(3, members.count())
        self.assertEqual("여신강림", title)
        self.assertEqual("checker", role)

    # def test_edit_project_members(self) -> None:
    #     # Given
    #     manager = self.create_user("manager", True)
    #     author = self.create_user("author", False)
    #     checker = self.create_user("checker", False)

    #     project = Project.objects.create(category="webtoon", title="판타지", description="한일/웹툰/시즌1", uploader=manager)

    #     Project_Member.objects.create(project=project, member=manager, role="manager")
    #     Project_Member.objects.create(project=project, member=author, role="author")
    #     Project_Member.objects.create(project=project, member=checker, role="checker")

    #     # When
    #     deleted_cnt, _ = Project_Member.objects.filter(project=project, member=checker).delete()
    #     new_checker = self.create_user("new_checker", False)
    #     result2 = Project_Member.objects.create(project=project, member=new_checker, role="checker")
    #     members = project.members.all()
        
    #     # Then
    #     with self.assertRaises(Project_Member.DoesNotExist):
    #         Project_Member.objects.filter(project=project, member=checker).get()

    #     self.assertEqual(deleted_cnt, 1)
    #     self.assertEqual("new_checker", result2.member.username)
    #     self.assertEqual(members.count(), 3)

    def test_create_a_project_and_translation_data(self) -> None:
        # Given
        manager = self.create_user("manager", True)
        author = User.objects.create(username="author", password="1234", is_admin=False)
        checker = User.objects.create(username="checker", password="1234", is_admin=False)

        data = {
            "manager": manager.pk,
            "author": author.pk,
            "checker": checker.pk,
            "category": "project_category",
            "title": "project_title",
            "description": "project_description",
            "filename": "rawdatas/a240c3c0b81911ed806db1505abef5d6.xlsx"
        }

        # When
        # 프로젝트생성
        try:
            with transaction.atomic():
                project = Project.objects.create(category=data["category"], title=data["title"], description=data["description"], uploader=manager)

                error_obj = Project.objects.get(pk=9999)

                # Project_Member.objects.create(project=project, member=manager, role="manager")
                # Project_Member.objects.create(project=project, member=author, role="author")
                # Project_Member.objects.create(project=project, member=checker, role="checker")
        except Exception:
            print("raise ERROR")

        result = Project.objects.filter(pk=project.pk).exists()
        self.assertEqual(result, False)
        # self.assertEqual(result.title, data["title"])

            #s3파일 가져오기
            # s3 = boto3.client("s3")
            # bucket = settings.AWS_BUCKET
            # file_name = data["filename"]
            # data = s3.get_object(Bucket=bucket, Key=file_name)
            # contents = data["Body"].read()
            # wb = load_workbook(filename=(io.BytesIO(contents)), data_only=True)
            # # 시트 불러오기
            # sheets = wb.sheetnames
            # # 시트접근
            # ws = wb["Sheet1"]
            # for row in ws.iter_rows(min_row=2):
            #     if row[0].value == None and row[1].value == None:
            #         continue
            #     Translation.objects.create(projects=project, number=row[0].value, remark=row[1].value, raw_data=row[2].value, trans_data=row[3].value)
            #     # print(data)
            #     print(row[0].value, row[1].value, row[2].value, row[3].value)
    



