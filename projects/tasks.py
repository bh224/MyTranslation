from config import settings
from celery import shared_task
from projects.models import Project, Project_Member
import boto3
import io
from openpyxl import load_workbook
from trans.models import CheckTranslation, Translation
from users.models import User


@shared_task
def test_get_result():
    total = 0
    for i in range(1000):
        total += i
    return True


@shared_task
def save_translation_data(filename, author_pk, checker_pk, project_pk):
    # 번역 & 체크데이터 저장
    author = User.objects.get(pk=author_pk)
    checker = User.objects.get(pk=checker_pk)
    project = Project.objects.get(pk=project_pk)
    # S3파일 불러오기
    s3 = boto3.client("s3")
    bucket = settings.AWS_BUCKET
    file_name = f"rawdatas/{filename}.xlsx"
    data = s3.get_object(Bucket=bucket, Key=file_name)
    contents = data["Body"].read()
    wb = load_workbook(filename=(io.BytesIO(contents)), data_only=True)
    ws = wb["Sheet1"]

    # 번역&체크데이터저장
    for row in ws.iter_rows(min_row=2):
        if row[0].value == None and row[1].value == None:
            break
            # print("error")
            # raise ParseError("occured")
        translation = Translation.objects.create(
            project=project,
            num=row[0].value,
            remark=row[1].value,
            origin_data=row[2].value,
            author=author,
        )
        check = CheckTranslation.objects.create(
            project=project, translation=translation, checker=checker
        )

    return True
