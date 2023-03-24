from django.conf import settings
from django.db import transaction
from rest_framework.test import APITestCase
from projects.models import Project, Project_Member
from trans.models import CheckTranslation, Translation
from users.models import User
import MySQLdb
import boto3
import io
from openpyxl import Workbook, load_workbook


class TestProjectAPI(APITestCase):
    def user_logged_in(self, username, is_admin):
        user = User.objects.create(username=username, is_admin=is_admin)
        user.set_password("1234")
        user.save()
        self.client.force_login(user)
        return user
    
    # def test_get_projects(self) -> None:
    #     # Given
    #     manager = self.user_logged_in("manager", True)
    #     # author = self.user_logged_in("author", False)
    #     # checker = self.user_logged_in("checker", False)

    #     [
    #         Project.objects.create(
    #             category="webtoon",
    #             title=f"title{i}",
    #             description= "한일/네이버웹툰",
    #             uploader=manager,
    #         )
    #         for i in range(1, 11)
    #     ]

    #     # When
    #     response = self.client.get(
    #         "/api/v1/projects/", {"page": 1},
    #     )

    #     result = response.json()

    #     # Then
    #     self.assertEqual(result[0]["title"], "title1")
    #     self.assertEqual(len(result), 10)


    # def test_create_a_project(self) -> None:
    #     # Given
    #     manager = self.user_logged_in("manager", True)
    #     author = self.user_logged_in("author", False)
    #     checker = self.user_logged_in("checker", False)

    #     # When
    #     response = self.client.post(
    #         "/api/v1/projects/",
    #         data = {
    #             "category": "webtoon",
    #             "title": "여신강림",
    #             "description": "한일/네이버웹툰",
    #             "manager": manager.pk,
    #             "author": author.pk,
    #             "checker": checker.pk,
    #         }
    #     )

    #     result = response.json()
    #     project_pk = result["pk"]
    #     project = Project_Member.objects.filter(project=project_pk)
    #     project_author = Project_Member.objects.filter(project=project_pk, member=author.pk).get()

    #     # Then
    #     self.assertEqual(result["title"], "여신강림")
    #     self.assertEqual(len(project), 3)
    #     self.assertEqual(project_author.member.username, "author")


    def test_upload_project_with_translation_data(self) -> None:
        # Given
        manager = self.user_logged_in("manager", True)
        author = User.objects.create(username="author", password="1234", is_admin=False)
        checker = User.objects.create(username="checker", password="1234", is_admin=False)
        
        data = {
            "manager": manager.pk,
            "author": author.pk,
            "checker": checker.pk,
            "category": "project_category",
            "title": "project_title",
            "description": "project_description",
            "filename": "rawdatas/a240c3c0b81911ed806db1505abef5d6.xlsx"
        }
            
        # When
        try:
            with transaction.atomic():
                # 프로젝트 and 프로젝트 멤버 등록
                project = Project.objects.create(category=data["category"], title=data["title"], description=data["description"], uploader=manager)
                # print(">>>프로젝트id", project.pk)
                Project_Member.objects.create(project=project, member=manager, role="manager")
                Project_Member.objects.create(project=project, member=author, role="author")
                Project_Member.objects.create(project=project, member=checker, role="checker")

                # S3파일 불러오기
                s3 = boto3.client("s3")
                bucket = settings.AWS_BUCKET
                file_name = data["filename"]
                data = s3.get_object(Bucket=bucket, Key=file_name)
                contents = data["Body"].read()
                wb = load_workbook(filename=(io.BytesIO(contents)), data_only=True)
                ws = wb["Sheet1"]

                # 데이터읽기
                for row in ws.iter_rows(min_row=2):
                    if row[0].value == None and row[1].value == None:
                        continue
                    translation = Translation.objects.create(project=project, number=row[0].value, remark=row[1].value, raw_data=row[2].value, trans_data=row[3].value)
                    CheckTranslation.objects.create(project=project, translation=translation.pk)
        except Exception:
            # print("ERROR")
            error_result =Project.objects.get(pk=1)
            # print(error_result)


        # Then
        result_trans = Translation.objects.filter(project=project.pk)
        result_check = CheckTranslation.objects.filter(project=project.pk)

        self.assertEqual(len(result_trans), len(result_check))

