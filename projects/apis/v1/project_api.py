from ast import Param
from urllib import request
from django.core.mail.message import EmailMessage
from django.db import transaction
from django.conf import settings
import boto3
import io
from openpyxl import Workbook, load_workbook
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.exceptions import ParseError
from comments.services.comment_service import get_comments_list
from projects.serializers import ProjectListSerializer, ProjectDetailSerializer, ProjectMembersSerializer
from comments.serializers import CommentSerializer
from projects.models import Project, Project_Member
from projects.tasks import save_translation_data, test_get_result
from trans.models import CheckTranslation, Translation
from users.models import User
from celery.result import AsyncResult
from config.celery import app
import time


class Tests(APIView):
    def get(self, request):
        task_id = test_get_result.delay()
        #print(task_id)
        result = AsyncResult(id=str(task_id), app=app)
        while True:
            time.sleep(1)
            if result.status == "SUCCESS":
                #print(result.status)
                result.forget()
            break
        return Response({"msg":"done"})

class Projects(APIView):
    """ 프로젝트 관련 """
    # 모든 프로젝트 불러오기
    def get(self, request):
        projects = Project.objects.all()
        # todo 페이지네이션
        serializer = ProjectListSerializer(projects, many=True)
        return Response(serializer.data)

    # 프로젝트 생성 & 프로젝트 멤버 추가 & 번역/체크 데이터 저장
    def post(self, request):
        manager_pk = request.data['manager']
        author_pk = request.data['author']
        checker_pk = request.data['checker']
        filename = request.data['filename']
        try:
            with transaction.atomic():
                # 프로젝트 추가
                serializer = ProjectDetailSerializer(data=request.data)
                if serializer.is_valid():
                    if request.user:
                        project = serializer.save(uploader=request.user)
                    else:
                        uploader = User.objects.get(pk=manager_pk)
                        project = serializer.save(uploader=uploader)
                else:
                    return Response(serializer.errors)
                # 프로젝트 멤버 추가
                members = [
                    {
                        "project": project.pk,
                        "member": manager_pk,
                        "role": "manager"
                    },
                    {
                        "project": project.pk,
                        "member": author_pk,
                        "role": "author"
                    },
                    {
                        "project": project.pk,
                        "member": checker_pk,
                        "role": "checker"
                    }
                ]
                members_serializer = ProjectMembersSerializer(data=members, many=True)
                if members_serializer.is_valid():
                    members_serializer.save()
                else:
                    raise ParseError("멤버추가 실패")
                
                save_translation_data.delay(filename, author_pk, checker_pk, project.pk)

                # # 번역 & 체크데이터 저장
                # author = User.objects.get(pk=author_pk)
                # checker = User.objects.get(pk=checker_pk)
                # # S3파일 불러오기
                # s3 = boto3.client("s3")
                # bucket = settings.AWS_BUCKET
                # file_name = f"rawdatas/{filename}.xlsx"
                # data = s3.get_object(Bucket=bucket, Key=file_name)
                # contents = data["Body"].read()
                # wb = load_workbook(filename=(io.BytesIO(contents)), data_only=True)
                # ws = wb["Sheet1"]

                # # 번역&체크데이터저장
                # for row in ws.iter_rows(min_row=2):
                #     if row[0].value == None and row[1].value == None:
                #         break
                #         # print("error")
                #         # raise ParseError("occured")
                #     translation = Translation.objects.create(project=project, num=row[0].value, remark=row[1].value, origin_data=row[2].value, author=author)
                #     check = CheckTranslation.objects.create(project=project, translation=translation, checker=checker)
                    
                # 생성된 프로젝트 반환
                serializer = ProjectDetailSerializer(project)
                return Response(serializer.data)

        except Exception:
            raise ParseError("Try Again! Error Occured")  

class ProjectMembers(APIView):
    def post(self, request):
        manager_pk = request.data['manager']
        author_pk = request.data['author']
        checker_pk = request.data['checker']

        manager = User.objects.get(pk=manager_pk)
        author = User.objects.get(pk=author_pk)
        checker = User.objects.get(pk=checker_pk)

        project = Project.objects.get(pk=2)

        try:
            with transaction.atomic():
                members = [
                    {
                        "project": 2,
                        "member": 1,
                        "role": "manager"
                    },
                    {
                        "project": 2,
                        "member": 1,
                        "role": "author"
                    }
                ]
                members_serializer = ProjectMembersSerializer(data=members, many=True)
                if members_serializer.is_valid():
                    saved_members = members_serializer.save()
                    serializer = ProjectMembersSerializer(saved_members, many=True)
                    return Response(serializer.data)
            return Response(serializer.errors)
        except Exception:
            raise ParseError("Error")

class ProjectDetail(APIView):
    def get(self, request, pk):
        project = Project.objects.get(pk=pk)
        serializer = ProjectDetailSerializer(project)
        return Response(serializer.data)

class SendEmail(APIView):
    from_email = settings.EMAIL_HOST_USER
    def get(self, request, pk):
        project = Project.objects.get(pk=pk)
        to_email = Project_Member.objects.get(project_id=pk, role="author").member.email
        subject = "이메일 테스트"
        to = [to_email]
        message = f"[{project.title}]의 체크가  완료되었습니다. 2차 수정 후 최종납품 해 주세요"
        EmailMessage(subject=subject, body=message, to=to, from_email=self.from_email).send()

        return Response({"message": "ok"})