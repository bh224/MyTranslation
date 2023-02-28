from django.shortcuts import render
from rest_framework.response import Response
from rest_framework.decorators import api_view
from openpyxl import Workbook, load_workbook
from projects.models import Project
from trans.models import Translation
import boto3
import io

# Create your views here.
@api_view(("GET",))
def excel_uploads(request):
    # 프로젝트
    # filename = request.get_params("file")
    project = Project.objects.get(pk=1)
    # S3에서 파일 가져오기
    s3 = boto3.client("s3")
    bucket = "mytranslation-manager"
    file_name = "rawdatas/test.xlsx"
    data = s3.get_object(Bucket=bucket, Key=file_name)
    contents = data["Body"].read()
    wb = load_workbook(filename=(io.BytesIO(contents)), data_only=True)
    # 파일불러오기
    wb = load_workbook(filename="test.xlsx")
    # 시트 불러오기
    sheets = wb.sheetnames
    # print(sheets) # 리스트로 반환

    # 시트접근
    ws1 = wb["Sheet1"]

    # 데이터접근
    for row in ws1.iter_rows(min_row=2):
        if row[0].value == None and row[1].value == None:
            continue
        data = Translation.objects.create(projects=project, number=row[0].value, remark=row[1].value, raw_data=row[2].value, trans_data=row[3].value)
        # print(row[0].value, row[1].value, row[2].value, row[3].value)
        print(data)
    return Response({"details":"done"})
